"""Tests for examples/readiness_demo/build_powerbi_export.py"""

from __future__ import annotations

import hashlib
import shutil
import sys
from datetime import date
from pathlib import Path

import openpyxl
import pytest

DEMO_DIR = Path(__file__).parent.parent
sys.path.insert(0, str(DEMO_DIR))

from build_powerbi_export import (
    CLIENT_CONTEXT_COLUMNS,
    POWERBI_EXPORT_COLUMNS,
    build_powerbi_export,
    normalize_answer,
    resolve_status,
    write_client_context,
    write_powerbi_export,
)

SAMPLE_WORKBOOK = DEMO_DIR / "client_intake_template.xlsx"


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _make_minimal_workbook(
    tmp_path: Path,
    *,
    with_req_map: bool = True,
    with_client_export: bool = True,
    preferred_markets: str | None = None,
    project_id: str | None = "TEST-001",
    decision_makers: str | None = "yes",
    initial_mw: float | None = 32.0,
) -> Path:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    if with_client_export:
        ws_ce = wb.create_sheet("Client_Export")
        ws_ce.append(
            ["assessment_date", "project_id", "preferred_markets",
             "decision_makers_identified", "initial_critical_it_mw"]
        )
        ws_ce.append(
            [date(2025, 9, 15), project_id, preferred_markets,
             decision_makers, initial_mw]
        )

    if with_req_map:
        ws_rm = wb.create_sheet("Requirement_Map")
        ws_rm.append(
            ["requirement_id", "requirement_name", "category", "severity",
             "intake_field", "answer_yes_status", "answer_no_status",
             "answer_partial_status", "default_status", "default_notes", "market_scope"]
        )
        ws_rm.append(
            ["REQ-001", "Phase 1 MW requirement defined", "capacity", "high",
             "initial_critical_it_mw", "complete", "open", "in_progress",
             "not_started", "No data", "NAM-NOVA"]
        )
        ws_rm.append(
            ["REQ-002", "Decision makers identified", "decision", "high",
             "decision_makers_identified", "complete", "open", "in_progress",
             "open", "Not confirmed", "NAM-NOVA"]
        )
        ws_rm.append(
            ["REQ-003", "No intake field req", "technical", "medium",
             "", "complete", "open", "in_progress",
             "complete", "Default notes here", "NAM-PHX"]
        )

    ws_pbi = wb.create_sheet("PowerBI_Export")
    ws_pbi.append(POWERBI_EXPORT_COLUMNS)

    out = tmp_path / "test_workbook.xlsx"
    wb.save(str(out))
    return out


def _file_md5(path: Path) -> str:
    return hashlib.md5(path.read_bytes()).hexdigest()


# ---------------------------------------------------------------------------
# normalize_answer
# ---------------------------------------------------------------------------

class TestNormalizeAnswer:
    def test_yes_values(self):
        for v in ("yes", "y", "true", "complete", "closed", "done", "1",
                  "Yes", "YES", "True", "Complete"):
            assert normalize_answer(v) == "yes", f"Expected 'yes' for {v!r}"

    def test_no_values(self):
        for v in ("no", "n", "false", "open", "not_started", "not started",
                  "No", "NO", "False"):
            assert normalize_answer(v) == "no", f"Expected 'no' for {v!r}"

    def test_partial_values(self):
        for v in ("partial", "in progress", "in_progress", "Partial", "In Progress"):
            assert normalize_answer(v) == "partial", f"Expected 'partial' for {v!r}"

    def test_blank_values(self):
        for v in (None, "", "0", "none", "n/a", "na", "N/A"):
            assert normalize_answer(v) == "blank", f"Expected 'blank' for {v!r}"

    def test_numeric_truthy_treated_as_yes(self):
        assert normalize_answer(32.0) == "yes"
        assert normalize_answer("32 MW") == "yes"
        assert normalize_answer("some text") == "yes"


# ---------------------------------------------------------------------------
# resolve_status
# ---------------------------------------------------------------------------

class TestResolveStatus:
    def _req(
        self,
        intake_field: str = "some_field",
        default_status: str = "open",
        default_notes: str = "default",
    ) -> dict:
        return {
            "intake_field": intake_field,
            "answer_yes_status": "complete",
            "answer_no_status": "not_started",
            "answer_partial_status": "in_progress",
            "default_status": default_status,
            "default_notes": default_notes,
        }

    def test_no_intake_field_returns_default(self):
        req = self._req(intake_field="", default_status="complete", default_notes="my notes")
        status, notes = resolve_status(req, None)
        assert status == "complete"
        assert notes == "my notes"

    def test_yes_value_returns_yes_status(self):
        req = self._req()
        status, _ = resolve_status(req, "complete")
        assert status == "complete"

    def test_no_value_returns_no_status(self):
        req = self._req()
        status, _ = resolve_status(req, "no")
        assert status == "not_started"

    def test_partial_value_returns_partial_status(self):
        req = self._req()
        status, _ = resolve_status(req, "in_progress")
        assert status == "in_progress"

    def test_blank_value_returns_default_status(self):
        req = self._req(default_status="open")
        status, _ = resolve_status(req, None)
        assert status == "open"

    def test_numeric_truthy_returns_yes_status(self):
        req = self._req()
        status, _ = resolve_status(req, 32.0)
        assert status == "complete"


# ---------------------------------------------------------------------------
# build_powerbi_export
# ---------------------------------------------------------------------------

class TestBuildPowerBIExport:
    def test_generates_correct_row_count(self, tmp_path):
        wb_path = _make_minimal_workbook(tmp_path)
        rows = build_powerbi_export(wb_path)
        assert len(rows) == 3

    def test_schema_columns_present(self, tmp_path):
        wb_path = _make_minimal_workbook(tmp_path)
        rows = build_powerbi_export(wb_path)
        for row in rows:
            for col in POWERBI_EXPORT_COLUMNS:
                assert col in row, f"Missing column: {col}"

    def test_intake_field_numeric_maps_to_yes_status(self, tmp_path):
        wb_path = _make_minimal_workbook(tmp_path, initial_mw=32.0)
        rows = build_powerbi_export(wb_path)
        capacity_row = next(r for r in rows if r["requirement_name"] == "Phase 1 MW requirement defined")
        assert capacity_row["status"] == "complete"

    def test_intake_field_yes_text_maps_to_yes_status(self, tmp_path):
        wb_path = _make_minimal_workbook(tmp_path, decision_makers="yes")
        rows = build_powerbi_export(wb_path)
        decision_row = next(r for r in rows if r["requirement_name"] == "Decision makers identified")
        assert decision_row["status"] == "complete"

    def test_intake_field_none_maps_to_default_status(self, tmp_path):
        wb_path = _make_minimal_workbook(tmp_path, initial_mw=None)
        rows = build_powerbi_export(wb_path)
        capacity_row = next(r for r in rows if r["requirement_name"] == "Phase 1 MW requirement defined")
        assert capacity_row["status"] == "not_started"

    def test_no_intake_field_uses_default_status(self, tmp_path):
        wb_path = _make_minimal_workbook(tmp_path)
        rows = build_powerbi_export(wb_path)
        technical_row = next(r for r in rows if r["requirement_name"] == "No intake field req")
        assert technical_row["status"] == "complete"
        assert technical_row["notes"] == "Default notes here"

    def test_market_from_client_export_single_value(self, tmp_path):
        wb_path = _make_minimal_workbook(tmp_path, preferred_markets="NAM-ATL")
        rows = build_powerbi_export(wb_path)
        for row in rows:
            assert row["market"] == "NAM-ATL", f"Expected NAM-ATL, got {row['market']!r}"

    def test_market_falls_back_to_requirement_map_scope_when_client_blank(self, tmp_path):
        wb_path = _make_minimal_workbook(tmp_path, preferred_markets=None)
        rows = build_powerbi_export(wb_path)
        capacity_row = next(r for r in rows if r["requirement_name"] == "Phase 1 MW requirement defined")
        assert capacity_row["market"] == "NAM-NOVA"
        technical_row = next(r for r in rows if r["requirement_name"] == "No intake field req")
        assert technical_row["market"] == "NAM-PHX"

    def test_market_with_comma_falls_back_to_requirement_map(self, tmp_path):
        wb_path = _make_minimal_workbook(tmp_path, preferred_markets="NAM-ATL, NAM-DFW")
        rows = build_powerbi_export(wb_path)
        capacity_row = next(r for r in rows if r["requirement_name"] == "Phase 1 MW requirement defined")
        assert capacity_row["market"] == "NAM-NOVA"

    def test_project_id_from_client_export(self, tmp_path):
        wb_path = _make_minimal_workbook(tmp_path, project_id="MYPROJ-001")
        rows = build_powerbi_export(wb_path)
        for row in rows:
            assert row["project_id"] == "MYPROJ-001"

    def test_project_id_defaults_to_unknown_when_blank(self, tmp_path):
        wb_path = _make_minimal_workbook(tmp_path, project_id=None)
        rows = build_powerbi_export(wb_path)
        for row in rows:
            assert row["project_id"] == "UNKNOWN"

    def test_date_from_client_export(self, tmp_path):
        wb_path = _make_minimal_workbook(tmp_path)
        rows = build_powerbi_export(wb_path)
        assert rows[0]["date"] == date(2025, 9, 15)

    def test_missing_requirement_map_raises_clearly(self, tmp_path):
        wb_path = _make_minimal_workbook(tmp_path, with_req_map=False)
        with pytest.raises(ValueError, match="Missing required sheet: Requirement_Map"):
            build_powerbi_export(wb_path)

    def test_missing_client_export_raises_clearly(self, tmp_path):
        wb_path = _make_minimal_workbook(tmp_path, with_client_export=False)
        with pytest.raises(ValueError, match="Missing required sheet: Client_Export"):
            build_powerbi_export(wb_path)


# ---------------------------------------------------------------------------
# write_powerbi_export + source integrity
# ---------------------------------------------------------------------------

class TestWritePowerBIExport:
    def test_generated_workbook_has_correct_header(self, tmp_path):
        wb_path = _make_minimal_workbook(tmp_path)
        rows = build_powerbi_export(wb_path)
        out_path = tmp_path / "output.xlsx"
        shutil.copy2(str(wb_path), str(out_path))
        write_powerbi_export(out_path, rows)

        wb = openpyxl.load_workbook(str(out_path), data_only=True)
        assert "PowerBI_Export" in wb.sheetnames
        ws = wb["PowerBI_Export"]
        header = [cell.value for cell in next(ws.iter_rows(max_row=1))]
        for col in POWERBI_EXPORT_COLUMNS:
            assert col in header, f"Missing column in header: {col}"

    def test_generated_workbook_row_count_matches(self, tmp_path):
        wb_path = _make_minimal_workbook(tmp_path)
        rows = build_powerbi_export(wb_path)
        out_path = tmp_path / "output.xlsx"
        shutil.copy2(str(wb_path), str(out_path))
        write_powerbi_export(out_path, rows)

        wb = openpyxl.load_workbook(str(out_path), data_only=True)
        ws = wb["PowerBI_Export"]
        data_rows = [
            r for r in ws.iter_rows(min_row=2, values_only=True)
            if any(v is not None for v in r)
        ]
        assert len(data_rows) == 3

    def test_source_workbook_not_modified(self, tmp_path):
        wb_path = _make_minimal_workbook(tmp_path)
        original_hash = _file_md5(wb_path)

        rows = build_powerbi_export(wb_path)
        out_path = tmp_path / "output.xlsx"
        shutil.copy2(str(wb_path), str(out_path))
        write_powerbi_export(out_path, rows)

        assert _file_md5(wb_path) == original_hash, "Source workbook was modified!"

    def test_pipeline_readable_columns(self, tmp_path):
        """Generated PowerBI_Export has the columns the pipeline expects."""
        wb_path = _make_minimal_workbook(tmp_path)
        rows = build_powerbi_export(wb_path)
        out_path = tmp_path / "pipeline_test.xlsx"
        shutil.copy2(str(wb_path), str(out_path))
        write_powerbi_export(out_path, rows)

        import polars as pl

        df = pl.read_excel(str(out_path), sheet_name="PowerBI_Export")
        for col in ("date", "status", "severity", "category", "market"):
            assert col in df.columns, f"Pipeline-required column missing: {col}"


# ---------------------------------------------------------------------------
# Integration: sample workbook
# ---------------------------------------------------------------------------

class TestSampleWorkbook:
    @pytest.mark.skipif(
        not SAMPLE_WORKBOOK.exists(),
        reason="Sample workbook not found",
    )
    def test_sample_workbook_generates_24_rows(self, tmp_path):
        rows = build_powerbi_export(SAMPLE_WORKBOOK)
        assert len(rows) == 24

    @pytest.mark.skipif(
        not SAMPLE_WORKBOOK.exists(),
        reason="Sample workbook not found",
    )
    def test_sample_workbook_source_unchanged(self, tmp_path):
        original_hash = _file_md5(SAMPLE_WORKBOOK)
        out_path = tmp_path / "generated.xlsx"

        rows = build_powerbi_export(SAMPLE_WORKBOOK)
        shutil.copy2(str(SAMPLE_WORKBOOK), str(out_path))
        write_powerbi_export(out_path, rows)

        assert _file_md5(SAMPLE_WORKBOOK) == original_hash, "Source workbook was modified!"

    @pytest.mark.skipif(
        not SAMPLE_WORKBOOK.exists(),
        reason="Sample workbook not found",
    )
    def test_sample_generated_workbook_readable_by_polars(self, tmp_path):
        import polars as pl

        out_path = tmp_path / "generated.xlsx"
        rows = build_powerbi_export(SAMPLE_WORKBOOK)
        shutil.copy2(str(SAMPLE_WORKBOOK), str(out_path))
        write_powerbi_export(out_path, rows)

        df = pl.read_excel(str(out_path), sheet_name="PowerBI_Export")
        assert len(df) == 24
        for col in POWERBI_EXPORT_COLUMNS:
            assert col in df.columns, f"Missing column: {col}"

    @pytest.mark.skipif(
        not SAMPLE_WORKBOOK.exists(),
        reason="Sample workbook not found",
    )
    def test_sample_statuses_match_defaults(self, tmp_path):
        """When Intake_Form is blank, all statuses come from Requirement_Map defaults."""
        rows = build_powerbi_export(SAMPLE_WORKBOOK)
        statuses = {r["requirement_name"]: r["status"] for r in rows}

        # Spot-check a few known defaults from Requirement_Map
        assert statuses["Phase 1 MW requirement defined"] == "complete"
        assert statuses["Phase 2 expansion scope"] == "in_progress"
        assert statuses["Minimum viable capacity confirmed"] == "not_started"
        assert statuses["Utility interconnection status confirmed"] == "open"
        assert statuses["Lease vs buy decision finalized"] == "not_started"
        assert statuses["Stakeholder alignment confirmed"] == "open"


# ---------------------------------------------------------------------------
# write_client_context
# ---------------------------------------------------------------------------

def _make_client_dict(
    *,
    project_id: str = "TEST-001",
    client_name: str = "Acme Corp",
    assessment_date=date(2025, 9, 15),
    target_decision_date=date(2026, 3, 1),
    initial_mw: float | None = 32.0,
) -> dict:
    return {
        "assessment_date": assessment_date,
        "project_id": project_id,
        "client_name": client_name,
        "project_name": "Project Falcon",
        "use_case": "AI training",
        "project_stage": "RFP",
        "target_decision_date": target_decision_date,
        "initial_critical_it_mw": initial_mw,
        "future_expansion_mw": 64.0,
        "preferred_markets": "NAM-NOVA",
        "commercial_model": "wholesale lease",
        "ready_for_cbre": "yes",
        "recommended_next_step": "Issue RFP",
    }


class TestWriteClientContext:
    def test_creates_file(self, tmp_path):
        out = tmp_path / "client_context.csv"
        write_client_context(_make_client_dict(), out)
        assert out.exists()

    def test_has_correct_columns(self, tmp_path):
        import csv as _csv

        out = tmp_path / "client_context.csv"
        write_client_context(_make_client_dict(), out)
        with open(out, newline="", encoding="utf-8") as f:
            reader = _csv.DictReader(f)
            assert list(reader.fieldnames) == CLIENT_CONTEXT_COLUMNS

    def test_has_exactly_one_data_row(self, tmp_path):
        import csv as _csv

        out = tmp_path / "client_context.csv"
        write_client_context(_make_client_dict(), out)
        with open(out, newline="", encoding="utf-8") as f:
            rows = list(_csv.DictReader(f))
        assert len(rows) == 1

    def test_project_id_written(self, tmp_path):
        import csv as _csv

        out = tmp_path / "client_context.csv"
        write_client_context(_make_client_dict(project_id="PROJ-XYZ"), out)
        with open(out, newline="", encoding="utf-8") as f:
            row = list(_csv.DictReader(f))[0]
        assert row["project_id"] == "PROJ-XYZ"

    def test_client_name_written(self, tmp_path):
        import csv as _csv

        out = tmp_path / "client_context.csv"
        write_client_context(_make_client_dict(client_name="NovaTech"), out)
        with open(out, newline="", encoding="utf-8") as f:
            row = list(_csv.DictReader(f))[0]
        assert row["client_name"] == "NovaTech"

    def test_assessment_date_formatted_as_iso(self, tmp_path):
        import csv as _csv

        out = tmp_path / "client_context.csv"
        write_client_context(_make_client_dict(assessment_date=date(2025, 9, 15)), out)
        with open(out, newline="", encoding="utf-8") as f:
            row = list(_csv.DictReader(f))[0]
        assert row["assessment_date"] == "2025-09-15"

    def test_target_decision_date_formatted_as_iso(self, tmp_path):
        import csv as _csv

        out = tmp_path / "client_context.csv"
        write_client_context(_make_client_dict(target_decision_date=date(2026, 3, 1)), out)
        with open(out, newline="", encoding="utf-8") as f:
            row = list(_csv.DictReader(f))[0]
        assert row["target_decision_date"] == "2026-03-01"

    def test_none_fields_written_as_empty_string(self, tmp_path):
        import csv as _csv

        client = _make_client_dict()
        client["client_name"] = None
        client["initial_critical_it_mw"] = None
        out = tmp_path / "client_context.csv"
        write_client_context(client, out)
        with open(out, newline="", encoding="utf-8") as f:
            row = list(_csv.DictReader(f))[0]
        assert row["client_name"] == ""
        assert row["initial_critical_it_mw"] == ""

    def test_zero_fields_written_as_empty_string(self, tmp_path):
        import csv as _csv

        client = _make_client_dict(initial_mw=0)
        out = tmp_path / "client_context.csv"
        write_client_context(client, out)
        with open(out, newline="", encoding="utf-8") as f:
            row = list(_csv.DictReader(f))[0]
        assert row["initial_critical_it_mw"] == ""

    def test_creates_parent_dirs(self, tmp_path):
        out = tmp_path / "deep" / "nested" / "client_context.csv"
        write_client_context(_make_client_dict(), out)
        assert out.exists()


class TestSampleWorkbookClientContext:
    @pytest.mark.skipif(
        not SAMPLE_WORKBOOK.exists(),
        reason="Sample workbook not found",
    )
    def test_sidecar_written_alongside_generated_workbook(self, tmp_path):
        import csv as _csv
        import openpyxl as _openpyxl

        out_wb = tmp_path / "generated.xlsx"
        ctx_out = tmp_path / "client_context.csv"

        shutil.copy2(str(SAMPLE_WORKBOOK), str(out_wb))
        wb = _openpyxl.load_workbook(str(SAMPLE_WORKBOOK), data_only=True)
        from build_powerbi_export import _read_client_export
        client = _read_client_export(wb)
        write_client_context(client, ctx_out)

        assert ctx_out.exists()
        with open(ctx_out, newline="", encoding="utf-8") as f:
            rows = list(_csv.DictReader(f))
        assert len(rows) == 1
        for col in CLIENT_CONTEXT_COLUMNS:
            assert col in rows[0], f"Missing column: {col}"
