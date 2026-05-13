"""
Readiness workbook builder.

Reads Requirement_Map + Client_Export sheets from a client intake workbook,
resolves each requirement's status from intake field answers, and produces:
  - a flat PowerBI_Export sheet (one row per requirement)
  - a client_context.csv sidecar (project metadata for downstream Power BI template)

The source workbook is never modified unless --in-place is passed to the CLI.
"""

from __future__ import annotations

import csv
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

import openpyxl

POWERBI_EXPORT_COLUMNS = [
    "date",
    "project_id",
    "category",
    "market",
    "requirement_name",
    "status",
    "severity",
    "notes",
]

CLIENT_CONTEXT_COLUMNS = [
    # Core project identity
    "assessment_date",
    "project_id",
    "client_name",
    "project_name",
    "project_type",
    # Requirement context
    "use_case",
    "primary_use_case",
    "project_stage",
    "readiness_stage",
    # Target specs
    "target_market",
    "target_region",
    "target_capacity_mw",
    "target_rack_density_kw",
    "target_live_date",
    "target_decision_date",
    "initial_critical_it_mw",
    "future_expansion_mw",
    # Commercial context
    "buyer_profile",
    "preferred_markets",
    "commercial_model",
    "ready_for_cbre",
    "recommended_next_step",
    # Report metadata
    "prepared_for",
    "prepared_by",
    "prepared_date",
    "version",
    "confidentiality",
    # Narrative
    "executive_summary",
    "key_decision_question",
]

REQUIREMENT_MAP_COLUMNS = [
    "requirement_id",
    "requirement_name",
    "category",
    "severity",
    "intake_field",
    "answer_yes_status",
    "answer_no_status",
    "answer_partial_status",
    "default_status",
    "default_notes",
    "market_scope",
]

# Keys whose values are formatted as ISO dates in client_context.csv.
_DATE_KEYS: frozenset[str] = frozenset(
    {"assessment_date", "target_decision_date", "target_live_date", "prepared_date"}
)

# Excel serial date base (Excel uses 1899-12-30 as epoch).
_EXCEL_EPOCH = date(1899, 12, 30)


def _from_excel_date(value: Any) -> date | None:
    """Convert an Excel serial date, datetime, or ISO string to a Python date."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)) and value > 0:
        return _EXCEL_EPOCH + timedelta(days=int(value))
    if isinstance(value, str):
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y", "%Y/%m/%d"):
            try:
                return datetime.strptime(value.strip(), fmt).date()
            except ValueError:
                continue
    return None


def normalize_answer(value: Any) -> str:
    """Return 'yes', 'no', 'partial', or 'blank' for a raw cell value.

    yes     → yes / y / true / complete / closed / done / 1
    no      → no / n / false / open / not_started / not started
    partial → partial / in progress / in_progress
    blank   → None / "" / 0 / none / n/a / na
    Any other non-empty string → 'yes' (field has content = answered)
    """
    if value is None:
        return "blank"
    s = str(value).strip().lower()
    if not s or s in ("0", "none", "n/a", "na"):
        return "blank"
    if s in ("yes", "y", "true", "complete", "closed", "done", "1"):
        return "yes"
    if s in ("no", "n", "false", "open", "not_started", "not started"):
        return "no"
    if s in ("partial", "in progress", "in_progress"):
        return "partial"
    return "yes"


def resolve_status(req: dict, intake_value: Any) -> tuple[str, str]:
    """Return (status, notes) for a Requirement_Map row given a raw intake value.

    If intake_field is blank, returns (default_status, default_notes).
    Otherwise normalizes intake_value and maps to the corresponding answer_*_status
    field; falls back to default_status on blank/unknown.
    """
    intake_field = (req.get("intake_field") or "").strip()
    default_status = req.get("default_status") or "open"
    default_notes = req.get("default_notes") or ""

    if not intake_field:
        return default_status, default_notes

    answer = normalize_answer(intake_value)
    if answer == "yes":
        notes = str(intake_value) if intake_value is not None else default_notes
        return req.get("answer_yes_status") or default_status, notes
    if answer == "no":
        return req.get("answer_no_status") or default_status, default_notes
    if answer == "partial":
        notes = str(intake_value) if intake_value is not None else default_notes
        return req.get("answer_partial_status") or default_status, notes
    # blank / unknown
    return default_status, default_notes


def _read_sheet_as_dicts(ws) -> list[dict]:
    """Read an openpyxl worksheet into a list of row dicts (row 1 = header)."""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    result = []
    for row in rows[1:]:
        if all(v is None for v in row):
            continue
        result.append(dict(zip(headers, row)))
    return result


def _read_client_export(wb: openpyxl.Workbook) -> dict:
    """Return Client_Export data as a dict from the header + first data row."""
    if "Client_Export" not in wb.sheetnames:
        raise ValueError("Missing required sheet: Client_Export")
    ws = wb["Client_Export"]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 1:
        raise ValueError("Client_Export is empty")
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    if len(rows) < 2:
        raise ValueError("Client_Export has no data row")
    values = rows[1]
    return dict(zip(headers, values))


def build_powerbi_export(workbook_path: Path) -> list[dict]:
    """Generate PowerBI_Export rows from Requirement_Map + Client_Export.

    Returns a list of row dicts with keys matching POWERBI_EXPORT_COLUMNS.
    Raises ValueError on missing sheets or missing required Requirement_Map fields.
    """
    wb = openpyxl.load_workbook(str(workbook_path), data_only=True)

    if "Requirement_Map" not in wb.sheetnames:
        raise ValueError("Missing required sheet: Requirement_Map")

    client = _read_client_export(wb)
    requirements = _read_sheet_as_dicts(wb["Requirement_Map"])

    if not requirements:
        raise ValueError("Requirement_Map has no data rows")

    raw_date = client.get("assessment_date")
    row_date = _from_excel_date(raw_date) or date.today()

    project_id = client.get("project_id")
    if not project_id or project_id == 0:
        project_id = "UNKNOWN"
    else:
        project_id = str(project_id).strip()

    client_market_raw = client.get("preferred_markets")
    if client_market_raw and str(client_market_raw).strip() and "," not in str(client_market_raw):
        client_market = str(client_market_raw).strip()
    else:
        client_market = None

    rows = []
    for idx, req in enumerate(requirements, start=2):
        for required_col in ("requirement_name", "category", "severity", "default_status"):
            if not req.get(required_col):
                raise ValueError(
                    f"Requirement_Map row {idx} missing required field: {required_col!r}"
                )

        intake_field = (req.get("intake_field") or "").strip()
        intake_value = client.get(intake_field) if intake_field else None

        status, notes = resolve_status(req, intake_value)
        market = client_market if client_market else (req.get("market_scope") or "")

        rows.append(
            {
                "date": row_date,
                "project_id": project_id,
                "category": str(req["category"]).strip(),
                "market": str(market).strip(),
                "requirement_name": str(req["requirement_name"]).strip(),
                "status": str(status).strip(),
                "severity": str(req["severity"]).strip(),
                "notes": str(notes).strip() if notes else "",
            }
        )

    return rows


def write_powerbi_export(output_path: Path, rows: list[dict]) -> None:
    """Write generated rows into the PowerBI_Export sheet of output_path.

    Clears all data rows (keeping header), then appends new rows.
    """
    wb = openpyxl.load_workbook(str(output_path))

    if "PowerBI_Export" not in wb.sheetnames:
        ws = wb.create_sheet("PowerBI_Export")
        ws.append(POWERBI_EXPORT_COLUMNS)
    else:
        ws = wb["PowerBI_Export"]
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row)

    for row in rows:
        ws.append([row[col] for col in POWERBI_EXPORT_COLUMNS])

    wb.save(str(output_path))


def write_client_export_sheet(output_path: Path, client: dict) -> None:
    """Overwrite the Client_Export sheet in output_path with the given client dict.

    Uses all keys present in client as headers so no intake field references are lost.
    The source workbook is not touched.
    """
    wb = openpyxl.load_workbook(str(output_path))
    if "Client_Export" in wb.sheetnames:
        del wb["Client_Export"]
    ws = wb.create_sheet("Client_Export")
    headers = list(client.keys())
    ws.append(headers)
    ws.append([client.get(h) for h in headers])
    wb.save(str(output_path))


def write_client_context(client: dict, output_path: Path) -> None:
    """Write a single-row client_context.csv from Client_Export data.

    Fields not present in client dict are written as empty strings.
    Date fields listed in _DATE_KEYS are formatted as YYYY-MM-DD.
    """
    output_path.parent.mkdir(parents=True, exist_ok=True)

    def _fmt(key: str) -> str:
        val = client.get(key)
        if val is None or val == 0:
            return ""
        if key in _DATE_KEYS:
            d = _from_excel_date(val)
            return d.isoformat() if d else str(val)
        return str(val).strip()

    row = {col: _fmt(col) for col in CLIENT_CONTEXT_COLUMNS}
    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=CLIENT_CONTEXT_COLUMNS)
        writer.writeheader()
        writer.writerow(row)
