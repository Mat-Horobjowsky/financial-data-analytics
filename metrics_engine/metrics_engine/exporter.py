import json
from pathlib import Path

import pandas as pd
from openpyxl.utils import get_column_letter

from metrics_engine.validator import ValidationReport


def export(
    long_metrics: pd.DataFrame,
    wide_metrics: pd.DataFrame,
    metric_dictionary: pd.DataFrame,
    validation_report: ValidationReport,
    output_dir: str | Path,
) -> None:
    out = Path(output_dir)
    out.mkdir(parents=True, exist_ok=True)

    long_metrics.to_csv(out / "long_metrics.csv", index=False)
    wide_metrics.to_csv(out / "wide_metrics.csv", index=False)
    metric_dictionary.to_csv(out / "metric_dictionary.csv", index=False)

    (out / "validation_report.json").write_text(
        json.dumps(
            {
                "status": validation_report.status,
                "errors": validation_report.errors,
                "warnings": validation_report.warnings,
            },
            indent=2,
        ),
        encoding="utf-8",
    )


def export_excel(
    long_metrics: pd.DataFrame,
    wide_metrics: pd.DataFrame,
    metric_dictionary: pd.DataFrame,
    validation_report: ValidationReport,
    output_dir: str | Path,
) -> None:
    out = Path(output_dir)
    out.mkdir(parents=True, exist_ok=True)

    report_df = pd.DataFrame({
        "key": (
            ["status"]
            + ["error"] * len(validation_report.errors)
            + ["warning"] * len(validation_report.warnings)
        ),
        "value": (
            [validation_report.status]
            + validation_report.errors
            + validation_report.warnings
        ),
    })

    sheets = [
        (long_metrics, "long_metrics"),
        (wide_metrics, "wide_metrics"),
        (metric_dictionary, "metric_dictionary"),
        (report_df, "validation_report"),
    ]

    with pd.ExcelWriter(out / "metrics_output.xlsx", engine="openpyxl") as writer:
        for df, name in sheets:
            df.to_excel(writer, sheet_name=name, index=False)
            ws = writer.sheets[name]
            ws.freeze_panes = "A2"
            for i, col in enumerate(ws.iter_cols(min_row=1, max_row=ws.max_row), start=1):
                max_len = max((len(str(cell.value or "")) for cell in col), default=0)
                ws.column_dimensions[get_column_letter(i)].width = min(max_len + 2, 60)
